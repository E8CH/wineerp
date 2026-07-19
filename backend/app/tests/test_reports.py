"""Story 5.1 — 종합 리포트 (FR10). manager 전용 + KST 일별 버킷."""
from __future__ import annotations

import uuid as _uuid
from collections.abc import Iterator
from datetime import UTC, date, datetime

import pytest
from fastapi.testclient import TestClient
from sqlalchemy.pool import StaticPool
from sqlmodel import Session, SQLModel, create_engine, select

from app.core.db import get_session
from app.main import app
from app.models.receiving import ReceivingRecord, ReceivingSource
from app.models.user import User, UserRole
from app.seed.wines import seed_demo_wines

API = "/api/v1"


@pytest.fixture
def engine():
    eng = create_engine(
        "sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool
    )
    SQLModel.metadata.create_all(eng)
    with Session(eng) as s:
        seed_demo_wines(s)
    return eng


@pytest.fixture
def client(engine) -> Iterator[TestClient]:
    def _session() -> Iterator[Session]:
        with Session(engine) as s:
            yield s

    app.dependency_overrides[get_session] = _session
    with TestClient(app) as c:
        yield c
    app.dependency_overrides.clear()


def _token(client, email="staff@wineerp.co") -> str:
    client.post(f"{API}/auth/signup", json={"email": email, "password": "pw123456"})
    return client.post(
        f"{API}/auth/login", data={"username": email, "password": "pw123456"}
    ).json()["access_token"]


def _manager(client, engine) -> str:
    token = _token(client, "mgr@wineerp.co")
    with Session(engine) as s:
        user = s.exec(select(User).where(User.email == "mgr@wineerp.co")).one()
        user.role = UserRole.manager
        s.add(user)
        s.commit()
    return token


def _h(t):
    return {"Authorization": f"Bearer {t}"}


def _vids(client, token) -> list[str]:
    return [
        v["id"]
        for v in client.post(
            f"{API}/scan", json={"code": "3760000000015"}, headers=_h(token)
        ).json()["products"][0]["vintages"]
    ]


def _staff_id(client, token) -> str:
    return client.get(f"{API}/auth/me", headers=_h(token)).json()["id"]


def _insert(engine, *, vid, sid, at, qty, **kw):
    with Session(engine) as s:
        s.add(
            ReceivingRecord(
                wine_vintage_id=_uuid.UUID(vid),
                staff_id=_uuid.UUID(sid),
                quantity=qty,
                received_at=at,
                **kw,
            )
        )
        s.commit()


def _report(client, token, **params):
    return client.get(f"{API}/reports/receiving", params=params, headers=_h(token))


# --- 권한 ---------------------------------------------------------------------


def test_staff_gets_403(client, engine):
    assert _report(client, _token(client), period="week").status_code == 403


def test_manager_gets_200(client, engine):
    assert _report(client, _manager(client, engine), period="week").status_code == 200


def test_requires_auth(client):
    assert client.get(f"{API}/reports/receiving").status_code == 401


def test_day_period_is_rejected(client, engine):
    """일간은 막대가 하나뿐이라 그래프의 의미가 없다."""
    assert _report(client, _manager(client, engine), period="day").status_code == 422


# --- 집계 ---------------------------------------------------------------------


def test_buckets_use_korean_calendar_day(client, engine):
    """🔴 UTC로 묶으면 오전 입고가 전날 막대로 간다."""
    staff = _token(client)
    vid, sid = _vids(client, staff)[0], _staff_id(client, staff)
    # KST 2026-07-15 08:00 == UTC 2026-07-14 23:00
    _insert(engine, vid=vid, sid=sid, at=datetime(2026, 7, 14, 23, 0, tzinfo=UTC), qty=5)

    mgr = _manager(client, engine)
    body = _report(client, mgr, period="month", anchor="2026-07-15").json()
    by_date = {b["date"]: b["quantity"] for b in body["buckets"]}
    assert by_date["2026-07-15"] == 5
    assert by_date["2026-07-14"] == 0


def test_empty_days_are_filled_with_zero(client, engine):
    """빼면 막대가 붙어 그려지고 "매일 들어왔다"로 읽힌다."""
    staff = _token(client)
    vid, sid = _vids(client, staff)[0], _staff_id(client, staff)
    _insert(engine, vid=vid, sid=sid, at=datetime(2026, 7, 15, 3, 0, tzinfo=UTC), qty=2)

    mgr = _manager(client, engine)
    body = _report(client, mgr, period="week", anchor="2026-07-15").json()
    assert len(body["buckets"]) == 7, "주간은 항상 7개 막대"
    assert sum(1 for b in body["buckets"] if b["quantity"] == 0) == 6


def test_month_has_one_bucket_per_day(client, engine):
    mgr = _manager(client, engine)
    body = _report(client, mgr, period="month", anchor="2026-07-15").json()
    assert len(body["buckets"]) == 31  # 7월


def test_top_products_sorted_by_quantity(client, engine):
    staff = _token(client)
    vids, sid = _vids(client, staff), _staff_id(client, staff)
    at = datetime(2026, 7, 15, 3, 0, tzinfo=UTC)
    _insert(engine, vid=vids[0], sid=sid, at=at, qty=3)
    _insert(engine, vid=vids[1], sid=sid, at=at, qty=9)

    mgr = _manager(client, engine)
    body = _report(client, mgr, period="week", anchor="2026-07-15").json()
    # 같은 제품의 두 빈티지 → 제품 단위로 합산된다
    assert body["top_products"][0]["quantity"] == 12
    assert body["total_quantity"] == 12
    assert body["distinct_wines"] == 2
    assert body["record_count"] == 2


def test_initial_setup_is_included(client, engine):
    """재고에 반영되는 수량이 리포트에 없으면 관리자가 무엇을 믿을지 모른다."""
    staff = _token(client)
    vid, sid = _vids(client, staff)[0], _staff_id(client, staff)
    at = datetime(2026, 7, 15, 3, 0, tzinfo=UTC)
    _insert(engine, vid=vid, sid=sid, at=at, qty=4)
    _insert(
        engine, vid=vid, sid=sid, at=at, qty=6, source=ReceivingSource.initial_setup
    )

    mgr = _manager(client, engine)
    body = _report(client, mgr, period="week", anchor="2026-07-15").json()
    assert body["total_quantity"] == 10


def test_soft_deleted_excluded(client, engine):
    staff = _token(client)
    vid, sid = _vids(client, staff)[0], _staff_id(client, staff)
    at = datetime(2026, 7, 15, 3, 0, tzinfo=UTC)
    _insert(engine, vid=vid, sid=sid, at=at, qty=4)
    _insert(engine, vid=vid, sid=sid, at=at, qty=99, deleted_at=datetime.now(UTC))

    mgr = _manager(client, engine)
    body = _report(client, mgr, period="week", anchor="2026-07-15").json()
    assert body["total_quantity"] == 4


def test_empty_period_returns_zero_buckets_not_error(client, engine):
    mgr = _manager(client, engine)
    body = _report(client, mgr, period="week", anchor="2020-01-01").json()
    assert body["total_quantity"] == 0
    assert body["top_products"] == []
    assert all(b["quantity"] == 0 for b in body["buckets"])


# --- Story 5.2: 엑셀 다운로드 ---------------------------------------------------


def _xlsx(client, token, **params):
    return client.get(f"{API}/reports/receiving.xlsx", params=params, headers=_h(token))


def _sheet(resp):
    import io as _io

    from openpyxl import load_workbook

    return load_workbook(_io.BytesIO(resp.content)).active


def test_xlsx_requires_manager(client, engine):
    assert _xlsx(client, _token(client), period="week").status_code == 403
    assert _xlsx(client, _manager(client, engine), period="week").status_code == 200


def test_xlsx_requires_auth(client):
    assert client.get(f"{API}/reports/receiving.xlsx").status_code == 401


def test_xlsx_has_correct_content_type_and_filename(client, engine):
    resp = _xlsx(client, _manager(client, engine), period="week", anchor="2026-07-15")
    assert (
        resp.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    # ASCII 파일명 — 한글은 Content-Disposition에서 클라이언트마다 처리가 갈린다.
    disposition = resp.headers["content-disposition"]
    assert "wineerp-receiving-2026-07-13_2026-07-19.xlsx" in disposition
    assert disposition.isascii()


def test_xlsx_contains_headers_and_rows(client, engine):
    staff = _token(client)
    vid, sid = _vids(client, staff)[0], _staff_id(client, staff)
    _insert(
        engine, vid=vid, sid=sid, at=datetime(2026, 7, 15, 3, 0, tzinfo=UTC), qty=7
    )

    resp = _xlsx(client, _manager(client, engine), period="week", anchor="2026-07-15")
    ws = _sheet(resp)
    headers = [c.value for c in ws[1]]
    assert headers == ["입고일시", "모델명", "생산자", "빈티지", "수량", "담당자", "구분", "메모"]

    row = [c.value for c in ws[2]]
    assert row[1] == "Château Margaux"
    assert row[4] == 7
    assert row[6] == "입고"


def test_xlsx_datetime_is_kst_not_utc(client, engine):
    """🔴 Excel 셀에는 시간대가 없다 — UTC를 그대로 쓰면 오전 입고가 전날로 보인다."""
    staff = _token(client)
    vid, sid = _vids(client, staff)[0], _staff_id(client, staff)
    # KST 2026-07-15 08:00 == UTC 2026-07-14 23:00
    _insert(
        engine, vid=vid, sid=sid, at=datetime(2026, 7, 14, 23, 0, tzinfo=UTC), qty=1
    )

    resp = _xlsx(client, _manager(client, engine), period="week", anchor="2026-07-15")
    ws = _sheet(resp)
    assert ws.cell(row=2, column=1).value == "2026-07-15 08:00"


def test_xlsx_writes_nv_not_blank(client, engine):
    """빈칸이면 "빠뜨린 값"으로 읽힌다 — NV는 유효 상태다."""
    staff = _token(client)
    mgr = _manager(client, engine)
    nv_vid = client.post(
        f"{API}/scan", json={"code": "3185370000060"}, headers=_h(staff)
    ).json()["products"][0]["vintages"][0]["id"]
    _insert(
        engine,
        vid=nv_vid,
        sid=_staff_id(client, staff),
        at=datetime(2026, 7, 15, 3, 0, tzinfo=UTC),
        qty=2,
    )
    ws = _sheet(_xlsx(client, mgr, period="week", anchor="2026-07-15"))
    assert ws.cell(row=2, column=4).value == "NV"


def test_xlsx_respects_period_filter(client, engine):
    """파일은 선택된 기간 필터를 그대로 반영한다(에픽 AC)."""
    staff = _token(client)
    vid, sid = _vids(client, staff)[0], _staff_id(client, staff)
    _insert(engine, vid=vid, sid=sid, at=datetime(2026, 7, 15, 3, 0, tzinfo=UTC), qty=1)
    _insert(engine, vid=vid, sid=sid, at=datetime(2026, 6, 15, 3, 0, tzinfo=UTC), qty=1)

    mgr = _manager(client, engine)
    ws = _sheet(_xlsx(client, mgr, period="week", anchor="2026-07-15"))
    assert ws.max_row == 2, "다른 달 기록이 섞이면 보고서가 틀린다"


def test_xlsx_excludes_soft_deleted(client, engine):
    staff = _token(client)
    vid, sid = _vids(client, staff)[0], _staff_id(client, staff)
    at = datetime(2026, 7, 15, 3, 0, tzinfo=UTC)
    _insert(engine, vid=vid, sid=sid, at=at, qty=1)
    _insert(engine, vid=vid, sid=sid, at=at, qty=99, deleted_at=datetime.now(UTC))

    ws = _sheet(_xlsx(client, _manager(client, engine), period="week", anchor="2026-07-15"))
    assert ws.max_row == 2


def test_xlsx_empty_period_still_has_headers(client, engine):
    """빈 기간에도 파일은 열려야 한다 — 헤더만 있는 시트."""
    ws = _sheet(_xlsx(client, _manager(client, engine), period="week", anchor="2020-01-01"))
    assert ws.max_row == 1
    assert ws.cell(row=1, column=1).value == "입고일시"


def test_memo_starting_with_equals_is_not_a_formula(client, engine):
    """🔴 Excel은 '='로 시작하는 문자열을 수식으로 분류한다.

    메모는 직원 자유 입력이고 이 파일은 관리자가 열어 회장에게 전달한다.
    `=cmd|'/c calc'!A0`는 프로세스 실행을 묻고, `=SUM(`은 파일 전체를 복구 대상으로
    만들며, `=A1*1000`은 메모 칸에 가짜 숫자를 렌더한다.
    """
    staff = _token(client)
    vid, sid = _vids(client, staff)[0], _staff_id(client, staff)
    _insert(
        engine, vid=vid, sid=sid, at=datetime(2026, 7, 15, 3, 0, tzinfo=UTC), qty=1,
        memo="=cmd|'/c calc'!A0",
    )
    ws = _sheet(_xlsx(client, _manager(client, engine), period="week", anchor="2026-07-15"))
    cell = ws.cell(row=2, column=8)
    assert cell.data_type == "s", "수식으로 분류되면 안 된다"
    assert cell.value.startswith("'="), "선행 아포스트로피로 무력화한다"


def test_quantity_stays_a_number_cell(client, engine):
    """수량은 집계 대상이므로 문자열이 되면 안 된다."""
    staff = _token(client)
    vid, sid = _vids(client, staff)[0], _staff_id(client, staff)
    _insert(engine, vid=vid, sid=sid, at=datetime(2026, 7, 15, 3, 0, tzinfo=UTC), qty=42)
    ws = _sheet(_xlsx(client, _manager(client, engine), period="week", anchor="2026-07-15"))
    assert ws.cell(row=2, column=5).data_type == "n"
    assert ws.cell(row=2, column=5).value == 42


def test_control_character_in_memo_does_not_500_the_export(client, engine):
    """메모 한 건의 제어문자가 그 기간 엑셀 전체를 영구히 500으로 만들면 안 된다."""
    staff = _token(client)
    vid, sid = _vids(client, staff)[0], _staff_id(client, staff)
    _insert(
        engine, vid=vid, sid=sid, at=datetime(2026, 7, 15, 3, 0, tzinfo=UTC), qty=1,
        memo="파손\x0b확인",
    )
    resp = _xlsx(client, _manager(client, engine), period="week", anchor="2026-07-15")
    assert resp.status_code == 200
    assert _sheet(resp).cell(row=2, column=8).value == "파손확인"


def test_top_products_order_is_deterministic_on_ties(client, engine):
    """동점에서 producer까지 비교하지 않으면 방언마다 5위가 달라진다."""
    from app.core.timeframe import Period, period_bounds
    from app.crud import report as report_crud
    from app.crud import wine as wine_crud

    staff = _token(client)
    sid = _staff_id(client, staff)
    at = datetime(2026, 7, 15, 3, 0, tzinfo=UTC)
    with Session(engine) as s:
        for producer in ("B Estate", "A Estate"):
            p = wine_crud.create_product(session=s, producer=producer, model_name="Same Name")
            v = wine_crud.add_vintage(s, wine_product_id=p.id, vintage=2020)
            s.add(
                ReceivingRecord(
                    wine_vintage_id=v.id, staff_id=_uuid.UUID(sid), quantity=5, received_at=at
                )
            )
        s.commit()
        start, end = period_bounds(Period.week, date(2026, 7, 15))
        first = report_crud.receiving_report(s, start=start, end=end)["top_products"]
        second = report_crud.receiving_report(s, start=start, end=end)["top_products"]

    assert first == second
    assert [p["producer"] for p in first[:2]] == ["A Estate", "B Estate"]
