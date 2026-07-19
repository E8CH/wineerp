"""엑셀 리포트 생성 (FR11).

⚠️ CSV가 아니라 진짜 `.xlsx`인 이유: CSV를 한국어 Windows Excel에서 열면 cp949/UTF-8
불일치로 **와인명이 깨진다**(한국 로케일에서 가장 흔한 실패). BOM을 붙여도 구분자·따옴표
처리가 환경마다 다르다. `.xlsx`는 인코딩 모호성이 없다.

⚠️ 일시는 **KST로 변환해 쓴다**. Excel 셀에는 시간대 개념이 없어 UTC를 그대로 넣으면
오전 입고가 전날로 보인다 — 4.1과 같은 함정이다.
"""
from __future__ import annotations

import io
from datetime import UTC, datetime, timedelta
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.core.config import settings

HEADERS = ["입고일시", "모델명", "생산자", "빈티지", "수량", "담당자", "구분", "메모"]
_SOURCE_LABEL = {"receiving": "입고", "initial_setup": "초기 세팅"}

# Excel이 수식으로 해석하는 선두 문자. 메모는 직원 자유 입력이고 모델명은 LLM이
# 라벨에서 읽은 값이라 둘 다 통제 밖이다.
_FORMULA_LEADS = ("=", "+", "-", "@")

# openpyxl이 거부하는 제어문자. 메모 한 건에 섞이면 그 기간 엑셀 전체가 500이 된다.
_ILLEGAL = {c: None for c in range(0x20) if c not in (0x09, 0x0A, 0x0D)}


def _safe_text(value: str | None) -> str:
    """엑셀 셀에 넣어도 안전한 문자열로 만든다.

    ⚠️ 수식 주입 방어. openpyxl은 `=`로 시작하는 문자열을 **수식으로 분류**한다.
    `=cmd|'/c calc'!A0`는 관리자에게 프로세스 실행을 묻고, `=SUM(`은 파일 전체를
    "복구할까요?"로 만들며, `=A1*1000`은 메모 칸에 **가짜 숫자**를 렌더한다 —
    재고 문서에서 최악이다. 회장에게 첨부되는 파일이므로 선두를 무력화한다.
    """
    if not value:
        return ""
    text = str(value).translate(_ILLEGAL)
    if text.startswith(_FORMULA_LEADS):
        # 선행 아포스트로피는 Excel에서 "이건 텍스트"라는 표준 표기다.
        return "'" + text
    return text


def build_receiving_workbook(rows: list[tuple]) -> bytes:
    """`crud.receiving.list_records` 결과를 엑셀 바이트로."""
    tz = ZoneInfo(settings.TIMEZONE)
    wb = Workbook()
    ws = wb.active
    ws.title = "입고 내역"

    ws.append(HEADERS)
    header_fill = PatternFill("solid", fgColor="123E7C")  # 네이비
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for rec, vintage, product, user in rows:
        received = rec.received_at
        if received.tzinfo is None:  # SQLite는 오프셋을 저장하지 않는다
            received = received.replace(tzinfo=UTC)
        ws.append(
            [
                received.astimezone(tz).strftime("%Y-%m-%d %H:%M"),
                _safe_text(product.model_name),
                _safe_text(product.producer),
                # NV는 빈칸이 아니라 "NV"로 쓴다 — 빈칸이면 "빠뜨린 값"으로 읽힌다.
                vintage.vintage if vintage.vintage is not None else "NV",
                rec.quantity,  # 숫자 셀로 들어가야 한다(집계 대상)
                _safe_text(user.email),
                _SOURCE_LABEL.get(str(rec.source), str(rec.source)),
                _safe_text(rec.memo),
            ]
        )

    widths = [18, 28, 22, 10, 8, 24, 12, 40]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width
    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def filename_for(start: datetime, end: datetime) -> str:
    """ASCII 파일명만 쓴다.

    한글 파일명은 `Content-Disposition`에서 RFC 5987 인코딩이 필요하고 클라이언트마다
    처리가 갈린다. 파일 내용은 한국어이므로 파일명까지 한글일 이유가 없다.
    """
    tz = ZoneInfo(settings.TIMEZONE)
    s = start.astimezone(tz).date().isoformat()
    # end는 배타적이다. 그대로 쓰면 7/13~7/19 주간 파일이 이름으로 7/20을 주장한다.
    e = (end.astimezone(tz) - timedelta(seconds=1)).date().isoformat()
    return f"wineerp-receiving-{s}_{e}.xlsx"
